from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook, load_workbook
import os




file_path = 'test.xlsx'
def write_to_excel(filepath, data_list):
    # 检查文件是否存在
    if os.path.exists(filepath):
        # 如果文件存在，加载工作簿
        workbook = load_workbook(filepath)
    else:
        # 如果文件不存在，创建新的工作簿和工作表
        workbook = Workbook()
    # 获取活动工作表
    sheet = workbook.active
    # 写入数据，将列表的值作为一行插入
    sheet.append(data_list)
    # 保存文件
    workbook.save(filepath)



# 配置 ChromeOptions
chrome_options = Options()
chrome_options.add_experimental_option("prefs", {
    "download.default_directory": r"D:\test",
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
})

# 初始化 WebDriver
driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()

# 打开目标网站
driver.get('https://smes.xfusion.com/smes/cpwrpt/#/common/InstructionFiling')

# 等待页面加载完成
wait = WebDriverWait(driver, 20)
# row_list=['file_name','file_big_class','file_small_class','abstract','keyword','file_condition','dept','file_link','upload_date','update_date']

def row_message():
    for i in range(1, 11):
        # row_list = []
        file_name_xpath = f'//*[@id="pane-1"]/div/div[3]/div/div[2]/div[3]/table/tbody/tr[{i}]/td[10]/div/a'
        abstract_xpath = f'//*[@id="pane-1"]/div/div[3]/div/div[2]/div[3]/table/tbody/tr[{i}]/td[11]/div'
        key_word_01_xpath = f'//*[@id="pane-1"]/div/div[3]/div/div[2]/div[3]/table/tbody/tr[{i}]/td[15]/div'
        key_word_02_xpath = f'//*[@id="pane-1"]/div/div[3]/div/div[2]/div[3]/table/tbody/tr[{i}]/td[7]/div'
        upload_date_xpath = f'//*[@id="pane-1"]/div/div[3]/div/div[2]/div[3]/table/tbody/tr[{i}]/td[20]/div'
        # 点击下载链接
        global downloaded_count
        downloaded_count=+1
        file_name = driver.find_element(By.XPATH, file_name_xpath).text
        abstract =  driver.find_element(By.XPATH, abstract_xpath).text
        key_word_01= driver.find_element(By.XPATH, key_word_01_xpath).text
        key_word_02 = driver.find_element(By.XPATH, key_word_02_xpath).text
        upload_date = driver.find_element(By.XPATH, upload_date_xpath).text
        # 获取文件类型（假设文件类型在链接文本中）
        row_list=[file_name,'MES+',abstract,'有效','产品工程与导入部','https://smes.xfusion.com/smes/cpwrpt/#/common/InstructionFiling',upload_date,'2025/8/4']
        write_to_excel(file_path, row_list)
        print(f'归档信息：{file_name}')


try:
    # 登录操作
    username = wait.until(EC.presence_of_element_located((By.NAME, 'username')))
    password = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="password-item"]/div/div/input')))
    username.send_keys('p50011748')
    password.send_keys('Pan787127839!')
    time.sleep(20)
    
    # login_button = wait.until(EC.element_to_be_clickable((By.ID, 'login-btn')))
    # login_button.click()
    # time.sleep(5)
    # 进入下载页面
    download_page = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="pane-1"]/div/div[1]/div[3]/div/button[1]')))
    download_page.click()
    time.sleep(5)

    total_files = 847
    files_per_page = 10
    total_pages = (total_files + files_per_page - 1) // files_per_page

    current_page = 1
    downloaded_count = 0


    while current_page <= total_pages and downloaded_count < total_files:
        # 等待当前页面加载完成
        row_message()
        # time.sleep(10)
        # 点击下一页按钮
        if current_page < total_pages:
            next_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="pane-1"]/div/div[3]/div/div[3]/button[2]')))
            next_button.click()
            current_page += 1
            time.sleep(5)
        else:
            break
finally:
    # 关闭浏览器
    driver.quit()
